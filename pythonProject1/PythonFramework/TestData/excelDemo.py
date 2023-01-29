import openpyxl

book =openpyxl.load_workbook("C:\\Users\\leksh\\Downloads\\Book (2).xlsx")
sheet =book.active
Dict = {}
cell =sheet.cell(row=1, column=4)
print(cell.value)
sheet.cell(row=2, column=2).value = "Box"

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)

print(sheet.max_column)

cell2 = sheet.cell(row=3, column=3)
print(cell2.value)
print(sheet['A2'].value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row =i,column=1).value == 5:

        for j in range(2,sheet.max_column+1):

            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value

print(Dict)

for j in range (1, sheet.max_column+1):
    print(sheet.cell(row=1, column=j).value)






